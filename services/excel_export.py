"""Excel export service – generate reports from Q&A or document data."""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from schemas import QAResult

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def export_qa_history(qa_results: list[QAResult], title: str = "Q&A Report") -> bytes:
    """Export a list of Q&A results to an Excel file (bytes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    headers = ["#", "Question", "Answer", "Sources (File, Page)", "Timestamp"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(headers))

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    for i, result in enumerate(qa_results, 1):
        sources_str = "; ".join(
            f"{s['filename']} p.{s['page_number']}" for s in result.sources
        )
        row = [i, result.query, result.answer, sources_str, now]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_extracted_data(
    rows: list[dict], headers: list[str] | None = None, title: str = "Extracted Data"
) -> bytes:
    """Export arbitrary extracted data (list of dicts) to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    if not rows:
        ws.cell(row=1, column=1, value="No data")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if headers is None:
        headers = list(rows[0].keys())

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(headers))

    for i, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=i, column=col, value=row_data.get(header, ""))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(
            len(str(header)) + 5, 15
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
