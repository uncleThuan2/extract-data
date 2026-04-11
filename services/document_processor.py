"""Document text extraction and chunking service.

Supported file types:
- PDF (.pdf)
- Word (.docx)
- Excel (.xlsx, .xls)
- CSV (.csv)
- Plain text (.txt, .md, .rst, .log, .json, .xml, .yaml, .yml)
- HTML (.html, .htm)
- Rich Text (.rtf)
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
from pathlib import PurePath
from langchain.text_splitter import RecursiveCharacterTextSplitter

from config import settings
from schemas import DocumentChunk

logger = logging.getLogger(__name__)

# File types that are NOT allowed (images, video, audio, binaries)
_BLOCKED_EXTENSIONS = frozenset({
    # Images
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif",
    ".webp", ".svg", ".ico", ".heic", ".heif", ".raw",
    # Video
    ".mp4", ".avi", ".mov", ".mkv", ".wmv", ".flv", ".webm", ".m4v",
    # Audio
    ".mp3", ".wav", ".flac", ".ogg", ".aac", ".wma", ".m4a",
    # Binary / Archives
    ".zip", ".tar", ".gz", ".rar", ".7z", ".exe", ".dll", ".bin", ".iso",
})

# Explicitly supported text-based file types
SUPPORTED_EXTENSIONS = frozenset({
    ".pdf", ".txt", ".md", ".rst", ".log",
    ".csv", ".json", ".xml", ".yaml", ".yml",
    ".html", ".htm", ".rtf",
    ".docx", ".xlsx", ".xls",
})


def is_supported_file(filename: str) -> bool:
    """Check if a file extension is supported for text extraction."""
    ext = PurePath(filename).suffix.lower()
    if ext in _BLOCKED_EXTENSIONS:
        return False
    return ext in SUPPORTED_EXTENSIONS


def get_supported_extensions_str() -> str:
    """Return a human-readable string of supported extensions."""
    return ", ".join(sorted(SUPPORTED_EXTENSIONS))


# ---------------------------------------------------------------------------
# Extractors per file type
# ---------------------------------------------------------------------------

def _extract_pdf(data: bytes, filename: str) -> list[dict]:
    import pymupdf  # pymupdf (fitz) – 5-10x faster than pdfplumber

    pages = []
    with pymupdf.open(stream=data, filetype="pdf") as doc:
        for i, page in enumerate(doc):
            text = page.get_text() or ""
            if text.strip():
                pages.append({"text": text, "page_number": i + 1, "filename": filename})
    return pages


def _extract_docx(data: bytes, filename: str) -> list[dict]:
    from docx import Document

    doc = Document(io.BytesIO(data))
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not full_text.strip():
        return []
    return [{"text": full_text, "page_number": 1, "filename": filename}]


def _extract_xlsx(data: bytes, filename: str) -> list[dict]:
    import inspect
    from openpyxl import load_workbook
    from openpyxl.worksheet.properties import WorksheetProperties

    # Some Excel files contain attributes (e.g. 'synchVertical') not recognised
    # by this version of openpyxl.  Patch __init__ to silently drop unknowns.
    _orig_wp_init = WorksheetProperties.__init__
    _valid_wp_params = set(inspect.signature(_orig_wp_init).parameters) - {"self"}

    def _tolerant_wp_init(self, **kwargs):
        _orig_wp_init(self, **{k: v for k, v in kwargs.items() if k in _valid_wp_params})

    WorksheetProperties.__init__ = _tolerant_wp_init

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    pages = []
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cell_strs = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cell_strs)
            if line.strip(" |"):
                rows_text.append(line)
        if rows_text:
            pages.append({
                "text": f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text),
                "page_number": idx,
                "filename": filename,
            })

    WorksheetProperties.__init__ = _orig_wp_init  # restore after use
    return pages


def _extract_csv(data: bytes, filename: str) -> list[dict]:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_text = []
    for row in reader:
        line = " | ".join(row)
        if line.strip(" |"):
            rows_text.append(line)
    if not rows_text:
        return []
    return [{"text": "\n".join(rows_text), "page_number": 1, "filename": filename}]


def _extract_json(data: bytes, filename: str) -> list[dict]:
    text = data.decode("utf-8", errors="replace")
    # Validate it's JSON, then store as pretty-printed text
    parsed = json.loads(text)
    pretty = json.dumps(parsed, indent=2, ensure_ascii=False)
    return [{"text": pretty, "page_number": 1, "filename": filename}]


def _extract_html(data: bytes, filename: str) -> list[dict]:
    from html.parser import HTMLParser

    class _TextExtractor(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.parts: list[str] = []
            self._skip = False

        def handle_starttag(self, tag: str, attrs: list) -> None:
            self._skip = tag in ("script", "style")

        def handle_endtag(self, tag: str) -> None:
            if tag in ("script", "style"):
                self._skip = False

        def handle_data(self, data: str) -> None:
            if not self._skip and data.strip():
                self.parts.append(data.strip())

    html_text = data.decode("utf-8", errors="replace")
    extractor = _TextExtractor()
    extractor.feed(html_text)
    text = "\n".join(extractor.parts)
    if not text.strip():
        return []
    return [{"text": text, "page_number": 1, "filename": filename}]


def _extract_plain_text(data: bytes, filename: str) -> list[dict]:
    text = data.decode("utf-8", errors="replace")
    if not text.strip():
        return []
    return [{"text": text, "page_number": 1, "filename": filename}]


_EXTRACTOR_MAP: dict[str, callable] = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".xls": _extract_xlsx,
    ".csv": _extract_csv,
    ".json": _extract_json,
    ".html": _extract_html,
    ".htm": _extract_html,
    # All plain text formats
    ".txt": _extract_plain_text,
    ".md": _extract_plain_text,
    ".rst": _extract_plain_text,
    ".log": _extract_plain_text,
    ".xml": _extract_plain_text,
    ".yaml": _extract_plain_text,
    ".yml": _extract_plain_text,
    ".rtf": _extract_plain_text,
}


def extract_text(file_bytes: bytes, filename: str) -> list[dict]:
    """Extract text from a supported file type."""
    ext = PurePath(filename).suffix.lower()
    extractor = _EXTRACTOR_MAP.get(ext)
    if extractor is None:
        raise ValueError(
            f"Unsupported file type: {ext}. "
            f"Supported: {get_supported_extensions_str()}"
        )
    pages = extractor(file_bytes, filename)
    logger.info("Extracted %d section(s) from %s", len(pages), filename)
    return pages


def chunk_pages(pages: list[dict]) -> list[DocumentChunk]:
    """Split extracted pages into smaller chunks for embedding."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.CHUNK_SIZE,
        chunk_overlap=settings.CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    chunks: list[DocumentChunk] = []
    for page in pages:
        splits = splitter.split_text(page["text"])
        for j, split_text in enumerate(splits):
            chunk_id = hashlib.sha256(
                f"{page['filename']}:{page['page_number']}:{j}".encode()
            ).hexdigest()[:16]

            chunks.append(
                DocumentChunk(
                    text=split_text,
                    metadata={
                        "filename": page["filename"],
                        "page_number": page["page_number"],
                        "chunk_index": j,
                    },
                    chunk_id=chunk_id,
                )
            )

    logger.info("Created %d chunks from %d pages", len(chunks), len(pages))
    return chunks


def process_document(file_bytes: bytes, filename: str) -> list[DocumentChunk]:
    """Full pipeline: extract text → chunk. Supports multiple file types."""
    pages = extract_text(file_bytes, filename)
    return chunk_pages(pages)
